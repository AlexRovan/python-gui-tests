import pytest
import openpyxl
from fixture.aplication import Application
from model.group import Group
import os

project_dir = os.path.dirname(os.path.abspath(__file__))


@pytest.fixture(scope="session")
def app(request):
    fixture = Application("C:\\Users\\Alex\\Downloads\\FreeAddressBookPortable\\AddressBook.exe")
    request.addfinalizer(fixture.destroy)
    return fixture


def pytest_generate_tests(metafunc):
    for fixture in metafunc.fixturenames:
        if fixture.startswith("excel_"):
            testdata = load_from_excel(fixture[5:])
            metafunc.parametrize(fixture, testdata,ids=[repr(x) for x in testdata])


def load_from_excel(name):
    group = []
    name=name[1:]
    wb = openpyxl.load_workbook(project_dir+'\\Data\\%s.xlsx' % name)
    ws = wb.get_sheet_by_name(name)
    for i in range(ws.max_row):
        cell_name = ws['B%s' % (i+1)]
        group.append(Group(name=cell_name.value))
    return group
