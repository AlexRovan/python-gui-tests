from openpyxl import Workbook
import os
import string
import random


project_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def random_string(prefix,maxlen = 10):
    symbols = string.ascii_letters + string.digits
    return prefix + "".join([random.choice(symbols) for i in range(random.randrange(maxlen))])


wb = Workbook()
ws = wb.active
ws.title = "groups"

n=3

ws['A1'] = 'Count_Group'
ws['B1'] = 'Name_Group'

for i in range(n):
    ws['A%s' % (i+1)] = i
    ws['B%s' % (i+1)] = random_string("name_")

ws['A%s' % (n+1)] = n+1
ws['B%s' % (n+1)] = ""

wb.save(filename = project_dir + '\\Data\\groups.xlsx')
